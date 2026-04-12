"""
scripts/build_dataset.py
Build the consolidated multi-sheet Excel dataset.
Pulls fresh data from WHO GHO API + merges original source data.

Run:  venv/bin/python3 scripts/build_dataset.py
"""

import os
import sys
import json
import datetime
import requests
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paths ─────────────────────────────────────────────────────────
ROOT      = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RAW_FILE  = os.path.join(ROOT, 'data', 'raw', 'data M new.xlsx')
OUT_FILE  = os.path.join(ROOT, 'data', 'processed', 'measles_bangladesh_consolidated.xlsx')
LOG_FILE  = os.path.join(ROOT, 'data', 'processed', 'update_log.txt')

# ── WHO GHO API ────────────────────────────────────────────────────
GHO_BASE  = 'https://ghoapi.azureedge.net/api'
TIMEOUT   = 30

def gho_fetch(indicator: str, country: str = 'BGD', retries: int = 3) -> pd.DataFrame:
    """Fetch a WHO GHO indicator with retry; return tidy DataFrame or empty."""
    url = (f"{GHO_BASE}/{indicator}"
           f"?$filter=SpatialDim eq '{country}'"
           f"&$orderby=TimeDim asc")
    for attempt in range(1, retries + 1):
        try:
            r = requests.get(url, timeout=TIMEOUT)
            r.raise_for_status()
            rows = r.json().get('value', [])
            if not rows:
                print(f"  [WARN] No data for {indicator}")
                return pd.DataFrame(columns=['Year', 'Value'])
            df       = pd.DataFrame(rows)
            year_col = df['TimeDim'].astype(int)
            val_col  = pd.to_numeric(df['NumericValue'], errors='coerce')
            result   = pd.DataFrame({'Year': year_col, 'Value': val_col})
            return result.dropna(subset=['Year']).reset_index(drop=True)
        except Exception as exc:
            print(f"  [WARN] {indicator} attempt {attempt}/{retries}: {exc}")
            if attempt == retries:
                print(f"  [ERROR] {indicator}: giving up, will use fallback data")
                return pd.DataFrame(columns=['Year', 'Value'])

def _clean(df, valcol):
    """Select and rename the Value column; coerce to numeric."""
    if df.empty:
        return pd.DataFrame(columns=['Year', valcol])
    df = df.copy()
    df['Year']  = df['Year'].astype(int)
    df[valcol]  = pd.to_numeric(df['Value'], errors='coerce')
    return df[['Year', valcol]].dropna(subset=[valcol]).reset_index(drop=True)

# ── 1. FETCH FROM WHO GHO ─────────────────────────────────────────
print("Fetching WHO GHO data (retries=3)...")

df_cases_gho = _clean(gho_fetch('WHS3_62'), 'Reported_Cases_WHO')
if df_cases_gho.empty:
    print("  [FALLBACK] Using built-in WHO GHO cases data (1975-2024)")
    _builtin_cases = [
        (1975,3030),(1976,40171),(1977,22973),(1978,18576),(1979,31895),
        (1980,11077),(1981,10441),(1982,18116),(1983,9569),(1984,11851),
        (1985,11699),(1986,14352),(1987,13631),(1988,11471),(1989,27327),
        (1990,1705),(1991,9292),(1992,5768),(1993,5443),(1994,9571),
        (1995,4995),(1996,4929),(1997,10329),(1998,6522),(1999,5666),
        (2000,5098),(2001,4414),(2002,3484),(2003,4067),(2004,9743),
        (2005,25934),(2006,6192),(2007,2924),(2008,2660),(2009,718),
        (2010,788),(2011,5625),(2012,1986),(2013,237),(2014,289),
        (2015,240),(2016,972),(2017,4001),(2018,2263),(2019,5827),
        (2020,2410),(2021,203),(2022,311),(2023,281),(2024,247),
    ]
    df_cases_gho = pd.DataFrame(_builtin_cases, columns=['Year','Reported_Cases_WHO'])
print(f"  Cases: {len(df_cases_gho)} rows  ({df_cases_gho.Year.min()}–{df_cases_gho.Year.max()})")

df_mcv1_who = _clean(gho_fetch('WHS8_110'), 'MCV1_WHO_Pct')
if df_mcv1_who.empty:
    print("  [FALLBACK] Using built-in MCV1 WHO data")
    _builtin_mcv1 = [
        (2000,74),(2001,77),(2002,75),(2003,76),(2004,81),(2005,88),(2006,83),
        (2007,89),(2008,92),(2009,93),(2010,88),(2011,93),(2012,88),(2013,91),
        (2014,94),(2015,97),(2016,97),(2017,97),(2018,97),(2019,97),(2020,97),
        (2021,97),(2022,96),(2023,96),(2024,96),
    ]
    df_mcv1_who = pd.DataFrame(_builtin_mcv1, columns=['Year','MCV1_WHO_Pct'])
print(f"  MCV1 WHO: {len(df_mcv1_who)} rows  ({df_mcv1_who.Year.min()}–{df_mcv1_who.Year.max()})")

df_mcv2_who = _clean(gho_fetch('MCV2'), 'MCV2_WHO_Pct')
if df_mcv2_who.empty:
    print("  [FALLBACK] Using built-in MCV2 WHO data")
    _builtin_mcv2 = [
        (2012,41),(2013,82),(2014,89),(2015,90),(2016,92),(2017,93),
        (2018,93),(2019,93),(2020,93),(2021,93),(2022,93),(2023,93),(2024,93),
    ]
    df_mcv2_who = pd.DataFrame(_builtin_mcv2, columns=['Year','MCV2_WHO_Pct'])
else:
    df_mcv2_who = df_mcv2_who[df_mcv2_who['MCV2_WHO_Pct'] > 0]
print(f"  MCV2 WHO: {len(df_mcv2_who)} rows  ({df_mcv2_who.Year.min()}–{df_mcv2_who.Year.max()})")

# ── 2. LOAD ORIGINAL EXCEL DATA ───────────────────────────────────
print("\nLoading original Excel data...")
import openpyxl as _oxl
wb_raw = _oxl.load_workbook(RAW_FILE)

# -- Sheet 'data': historical multi-source coverage + cases --------
ws_d = wb_raw['data']
rows = list(ws_d.iter_rows(values_only=True))

MISSING = {'-', '--', '—', None, ''}

def to_num(v):
    if v in MISSING or (isinstance(v, str) and v.strip() in MISSING):
        return np.nan
    try:
        return float(v)
    except (ValueError, TypeError):
        return np.nan

hist_rows = []
for row in rows[2:]:          # skip 2-row header
    yr = row[0]
    if yr is None:
        continue
    try:
        yr = int(float(str(yr)))
    except (ValueError, TypeError):
        continue
    if not (1975 <= yr <= 2030):
        continue
    hist_rows.append({
        'Year'           : yr,
        'Infected_Orig'  : to_num(row[1]),
        'MCV1_Admin'     : to_num(row[2]),
        'MCV1_Official'  : to_num(row[3]),
        'MCV1_WHO_Orig'  : to_num(row[4]),
        'MCV2_Admin'     : to_num(row[5]),
        'MCV2_Official'  : to_num(row[6]),
        'MCV2_WHO_Orig'  : to_num(row[7]),
    })

df_orig = pd.DataFrame(hist_rows)
print(f"  Original sheet 'data': {len(df_orig)} rows  ({df_orig.Year.min()}–{df_orig.Year.max()})")

# -- Sheet '2026': outbreak district data --------------------------
ws_26 = wb_raw['2026']
rows26 = list(ws_26.iter_rows(values_only=True))

dist_rows = []
for row in rows26[2:62]:     # rows 2-61 (0-indexed)
    name = row[12]
    inf  = row[13]
    dth  = row[15]
    if name is None or str(name).strip() in MISSING:
        continue
    try:
        inf = int(float(inf)) if inf not in MISSING else 0
        dth = int(float(dth)) if dth not in MISSING else 0
    except (ValueError, TypeError):
        inf, dth = 0, 0
    dist_rows.append({
        'District'  : str(name).strip(),
        'Cases'     : inf,
        'Deaths'    : dth,
    })

df_dist = pd.DataFrame(dist_rows)

# Division mapping (Bangladesh 8 divisions → districts)
DIST_TO_DIV = {
    'Dhaka':'Dhaka','Gazipur':'Dhaka','Narayangang':'Dhaka','Narshindi':'Dhaka',
    'Manikgang':'Dhaka','Munshiganj':'Dhaka','Rajbari':'Dhaka','Kishorgang':'Dhaka',
    'Faidpur':'Dhaka','Madaripur':'Dhaka','Gopalgong':'Dhaka','Tangail':'Dhaka',
    'Shariyatpur':'Dhaka',
    'Rajshahi':'Rajshahi','Chapainabagonj':'Rajshahi','Naogan':'Rajshahi',
    'Bagura':'Rajshahi','Natore':'Rajshahi','Pabna':'Rajshahi','Sirajgang':'Rajshahi',
    'Joypurhat':'Rajshahi','Jaypurhat':'Rajshahi',
    'Chottogram':'Chattogram','Coxsbazar':'Chattogram','Cumilla':'Chattogram',
    'Chadpur':'Chattogram','Noakhali':'Chattogram','Feni':'Chattogram',
    'Laxmipur':'Chattogram','Bramanbaria':'Chattogram','Khagrachori':'Chattogram',
    'Rangamati':'Chattogram','Bandarban':'Chattogram',
    'Khulna':'Khulna','Jessore':'Khulna','Jassore':'Khulna','Satkhira':'Khulna','Norail':'Khulna',
    'Jhinadah':'Khulna','Magura':'Khulna','Kustia':'Khulna','Chuyadanga':'Khulna',
    'Meherpur':'Khulna','Bagarhat':'Khulna',
    'Barisal':'Barisal','Potuakhali':'Barisal','Pirozpur':'Barisal',
    'Jhalkathi':'Barisal','Barguna':'Barisal','Bhola':'Barisal',
    'Syhlet':'Sylhet','Moulovibazar':'Sylhet','Habigang':'Sylhet','Sunamgang':'Sylhet',
    'Rangpur':'Rangpur','Nilfamari':'Rangpur','Gaibanda':'Rangpur','Kurigram':'Rangpur',
    'Lalmanirhat':'Rangpur','Dinajpur':'Rangpur','Thakurgaon':'Rangpur',
    'Panchagar':'Rangpur',
    'Maymansingh':'Mymensingh','Netrokona':'Mymensingh','Jamalpur':'Mymensingh',
    'Sherpur':'Mymensingh',
}
df_dist['Division'] = df_dist['District'].map(DIST_TO_DIV).fillna('Unknown')

total_hosp = df_dist['Cases'].sum()
df_dist['Pct_of_total'] = (df_dist['Cases'] / total_hosp * 100).round(2)
df_dist['CFR_pct']      = np.where(
    df_dist['Cases'] > 0,
    (df_dist['Deaths'] / df_dist['Cases'] * 100).round(2), 0.0)
df_dist = df_dist.sort_values('Cases', ascending=False).reset_index(drop=True)

# Division-level aggregation
df_div = df_dist.groupby('Division', as_index=False).agg(
    Districts_affected = ('District', 'count'),
    Cases              = ('Cases', 'sum'),
    Deaths             = ('Deaths', 'sum'),
)
df_div['Pct_of_total'] = (df_div['Cases'] / total_hosp * 100).round(2)
df_div['CFR_pct']      = np.where(
    df_div['Cases'] > 0,
    (df_div['Deaths'] / df_div['Cases'] * 100).round(2), 0.0)
df_div = df_div.sort_values('Cases', ascending=False).reset_index(drop=True)

print(f"  Outbreak districts (2026): {len(df_dist)} rows")
print(f"  Outbreak divisions (2026): {len(df_div)} rows")

# ── 3. PREPARE FULL CASE SERIES ───────────────────────────────────
print("\nPreparing case series...")

case_full = df_cases_gho.copy()
# Append placeholder row for 2026 (not yet WHO-reported)
yr2026 = pd.DataFrame([{'Year': 2026, 'Reported_Cases_WHO': np.nan}])
case_full = pd.concat([case_full, yr2026], ignore_index=True)
print(f"  Case series: {len(case_full)} rows  ({int(case_full.Year.min())}–{int(case_full.Year.max())})")

# ── 4. WRITE MULTI-SHEET EXCEL ────────────────────────────────────
print(f"\nWriting consolidated dataset → {OUT_FILE}")

with pd.ExcelWriter(OUT_FILE, engine='openpyxl') as writer:

    # ── Sheet 1: cases_full (1975-2024, WHO GHO) ──────────────────
    cf = case_full[case_full['Year'] <= 2024].copy()
    cf.columns = ['Year', 'Reported_Cases_WHO']
    cf['Note'] = ''
    cf.loc[cf['Year'] == 2012, 'Note'] = 'MCV2 introduced'
    cf.loc[cf['Year'] == 2020, 'Note'] = 'COVID-19 immunisation disruption'
    cf.to_excel(writer, sheet_name='cases_full', index=False)

    # ── Sheet 2: mcv1_coverage ────────────────────────────────────
    mc1 = df_orig[df_orig['Year'].between(2000, 2025)][
        ['Year','MCV1_WHO_Orig','MCV1_Official','MCV1_Admin']].copy()
    mc1 = mc1.merge(df_mcv1_who.rename(columns={'MCV1_WHO_Pct':'MCV1_WHO_GHO'}),
                    on='Year', how='outer').sort_values('Year')
    mc1.columns = ['Year','MCV1_WHO_OrigSource','MCV1_Official_%',
                   'MCV1_Admin_%','MCV1_WHO_GHO_%']
    mc1.to_excel(writer, sheet_name='mcv1_coverage', index=False)

    # ── Sheet 3: mcv2_coverage ────────────────────────────────────
    mc2 = df_orig[df_orig['Year'].between(2012, 2025)][
        ['Year','MCV2_WHO_Orig','MCV2_Official','MCV2_Admin']].copy()
    mc2 = mc2.merge(df_mcv2_who.rename(columns={'MCV2_WHO_Pct':'MCV2_WHO_GHO'}),
                    on='Year', how='outer').sort_values('Year')
    mc2.columns = ['Year','MCV2_WHO_OrigSource','MCV2_Official_%',
                   'MCV2_Admin_%','MCV2_WHO_GHO_%']
    mc2.to_excel(writer, sheet_name='mcv2_coverage', index=False)

    # ── Sheet 4: outbreak_2026 ────────────────────────────────────
    outbreak_meta = pd.DataFrame([
        ['Data type',          'Hospital-based surveillance'],
        ['Report date',        '2026-04-09'],
        ['Suspected cases',    12320],
        ['Lab-confirmed',      2241],
        ['Hospitalised',       6883],
        ['Deaths',             166],
        ['CFR (% of suspected)', round(166/12320*100, 2)],
        ['% deaths in 0-5 yr', 71.0],
        ['% deaths in 5-18 yr',12.0],
        ['% deaths in adults', 17.0],
        ['% deaths in children (0-18)', 92.86],
        ['Districts affected', 61],
        ['Source',             'DGHS / Hospital sentinel surveillance'],
    ], columns=['Metric', 'Value'])
    outbreak_meta.to_excel(writer, sheet_name='outbreak_2026_summary', index=False)
    df_div[['Division','Districts_affected','Cases','Deaths',
            'Pct_of_total','CFR_pct']].to_excel(
        writer, sheet_name='outbreak_2026_divisions', index=False)
    df_dist[['Division','District','Cases','Deaths',
             'Pct_of_total','CFR_pct']].to_excel(
        writer, sheet_name='outbreak_2026_districts', index=False)

    # ── Sheet 7: update_log ───────────────────────────────────────
    log_df = pd.DataFrame([{
        'Updated': datetime.datetime.now().isoformat(timespec='seconds'),
        'Source_Cases'  : 'WHO GHO WHS3_62',
        'Source_MCV1'   : 'WHO GHO WHS8_110',
        'Source_MCV2'   : 'WHO GHO MCV2',
        'Source_2026'   : 'DGHS hospital surveillance (manual)',
        'Cases_rows'    : len(df_cases_gho),
        'MCV1_rows'     : len(df_mcv1_who),
        'MCV2_rows'     : len(df_mcv2_who),
    }])
    log_df.to_excel(writer, sheet_name='update_log', index=False)

# ── Basic styling ─────────────────────────────────────────────────
wb = openpyxl.load_workbook(OUT_FILE)
HDR = PatternFill('solid', fgColor='1F4E79')
HDR_FONT = Font(bold=True, color='FFFFFF', size=11)
BORDER = Border(
    bottom=Side(style='thin', color='AAAAAA'),
    right=Side(style='thin', color='DDDDDD'),
)

for ws in wb.worksheets:
    for cell in ws[1]:          # header row
        cell.fill = HDR
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 28
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 30)
    ws.freeze_panes = 'A2'

wb.save(OUT_FILE)

# ── Append to log file ────────────────────────────────────────────
with open(LOG_FILE, 'a') as f:
    f.write(f"{datetime.datetime.now().isoformat(timespec='seconds')} — dataset rebuilt OK\n")

print(f"\n✓ Done: {OUT_FILE}")
print(f"  Sheets: {[ws.title for ws in wb.worksheets]}")
